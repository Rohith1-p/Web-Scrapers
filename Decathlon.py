def get_payload(category_url,page,cat_id):
    
    print("@@@@@@@@@@@@@@@in the get_payload ", cat_id, "Page ",page)
    
    if 'query' not in category_url:
        print("####################### in category page ########################")



        url = "https://th8ix1g260-dsn.algolia.net/1/indexes/*/queries?x-algolia-agent=Algolia%20for%20JavaScript%20(3.35.1)%3B%20Browser%20(lite)&x-algolia-application-id=TH8IX1G260&x-algolia-api-key=cd4f09aea8452f426737260804316800"


        payload = json.dumps({
          "requests": [
            {
              "indexName": "India_prod_dosav2_category",
              "params": "analyticsTags=%5B%22Web%22%2C%22Desktop%22%5D&clickAnalytics=true&facets=%5B%22tes_gender_en%22%2C%22size%22%2C%22tes_415_en%22%2C%22tes_sold_by_en%22%2C%22tes_product_nature_en%22%2C%22tes_sport_en%22%2C%22tes_brand_en%22%2C%22tes_color_en%22%2C%22price%22%5D&highlightPostTag=%3C%2Fais-highlight-0000000000%3E&highlightPreTag=%3Cais-highlight-0000000000%3E&maxValuesPerFacet=100&page="+str(page)+"&query="+str(cat_id)+"&tagFilters="
            }
          ]
        })


#         response = requests.request("POST", url, headers=headers, data=payload)

#         print(response.text)
    else:
        print("################## in the query search #######################")
        

        url = "https://th8ix1g260-dsn.algolia.net/1/indexes/*/queries?x-algolia-agent=Algolia%20for%20JavaScript%20(3.35.1)%3B%20Browser%20(lite)&x-algolia-application-id=TH8IX1G260&x-algolia-api-key=cd4f09aea8452f426737260804316800"

        payload = json.dumps({
          "requests": [
            {
              "indexName": "India_prod_dosav2",
              "params": "analyticsTags=%5B%22Web%22%2C%22Desktop%22%5D&clickAnalytics=true&facets=%5B%22gender_id_en%22%2C%22size%22%2C%22price%22%2C%22clearance_id%22%2C%22sport_en%22%2C%22manufacturer_en%22%2C%22nature_id_en%22%2C%22business_color%22%5D&highlightPostTag=%3C%2Fais-highlight-0000000000%3E&highlightPreTag=%3Cais-highlight-0000000000%3E&maxValuesPerFacet=100&page="+str(page)+"&query="+cat_id+"&tagFilters="
            }
          ]
        })
        
    return payload
        



import requests
import json
from scrapy.selector import Selector
import requests
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import datetime
import time
from datetime import datetime

sheet_flag = False

timestamp = time.time()


time_now = datetime.now().strftime('%F_%H_%M_%S')
workbook_name = time_now +'Decathlon.xlsx'


file_name = "urls.xlsx"

df = pd.read_excel(file_name)
urls = list(df['Urls'])
page_nums = list(df['Page Numbers'])

data_dict = {'ProductTitle':[],'Price':[],'MRP':[], 'Colour':[], 'URL':[], 'Product_id':[], 'Size':[], 'Category':[],'Category Sub':[],'Category URL':[]}
for i in range(len(urls)):
    
    category_url = urls[i]
    
    print('Category URL ----> ',category_url)
    
    
    if 'query' not in category_url:
        
        try:
            category = ' '.join(category_url.split('?')[0].split('/')[-1].split('-')[:-1])
        except:
            category = '-'

        print('category',category)
        try:
            cat_id = category_url.split('?')[0].split('/')[-1].split('-')[-1]
        except:
            cat_id = '-'
    else:
        try:
            cat_id = category_url.split('=')[-1]
            print("cat_id --- >",cat_id)
        except:
            continue

        try:
            category = ' '.join(category_url.split('=')[-1].split('%20'))
        except:
            category = '-'
            
    for page in range(page_nums[i]):
    
        url = "https://th8ix1g260-dsn.algolia.net/1/indexes/*/queries?x-algolia-agent=Algolia%20for%20JavaScript%20(3.35.1)%3B%20Browser%20(lite)&x-algolia-application-id=TH8IX1G260&x-algolia-api-key=cd4f09aea8452f426737260804316800"

        headers = {
                  'Content-Type': 'application/json'
                }

        print("(((((((((((((before payload)))))))))))))")
        payload = get_payload(category_url,page,cat_id)
        print("{{{{{{{{{{{{{{after payload}}}}}}}}}}}}}}")
        print(payload)
        response = requests.request("POST", url, headers=headers, data=payload)
        time.sleep(0.1)
        print(response.text)



        data = json.loads(response.text)
    #         print("data",data)

        values = data['results'][0]['hits']
        print("values ---->#########",values)

        for i in values:
            print('in values',i)
            product_name = i['name_en']
            discount_price = i['price']
            actual_price = i['price_mrp']
            Category_sub = i['category_en']
            rating = i['rating']
            url = i['link_en']

            product_id = i['objectID']
            try:
                color = i['business_color']
            except:
                color ='-'

            model_id = i['model_id']

            proudct_url = 'https://www.decathlon.in/p/'+ str(model_id)+'/'+ url
            print(proudct_url)


            sizes_variation = i['variations']['modelArr'][0]['itemArr']

            sizes_var = []
            for sizes in sizes_variation:
                sizes_var.append(sizes['item']['size'])



            data_dict['ProductTitle'].append(product_name)
            data_dict['Price'].append(discount_price)
            data_dict['MRP'].append(actual_price)
            data_dict['Colour'].append(color)
            data_dict['URL'].append(proudct_url)
            data_dict['Product_id'].append(product_id)
            data_dict['Size'].append(str(sizes_var))
            data_dict['Category'].append(category)
            data_dict['Category Sub'].append(Category_sub)
            data_dict['Category URL'].append(category_url)
        print(data_dict)


        data_df = pd.DataFrame(data_dict)
        data_df = data_df.astype(str)
        print(data_df)
        columns = data_df.columns.to_list()    
        df_values = data_df.values.tolist()
        df_values.insert(0,columns)

        print("sheet_flag",sheet_flag)
        if sheet_flag == False:

            sheet_flag = True
            wb = openpyxl.Workbook()
            ws = wb.active
            for row in df_values:
                ws.append(row)
                wb.save(workbook_name)

            data_dict = {'ProductTitle':[],'Price':[],'MRP':[], 'Colour':[], 'URL':[], 'Product_id':[], 'Size':[], 'Category':[],'Category Sub':[],'Category URL':[]}


        else:
            df_values = data_df.values.tolist()
            wb = load_workbook(workbook_name)
            page = wb.active
            for row in df_values:
                #print(row)
                page.append(row)
                wb.save(filename=workbook_name)
            data_dict = {'ProductTitle':[],'Price':[],'MRP':[], 'Colour':[], 'URL':[], 'Product_id':[], 'Size':[], 'Category':[],'Category Sub':[],'Category URL':[]}


